"""
Energy Model Benchmarking & QA/QC Platform
===========================================
Run:  streamlit run app.py
Deps: pip install streamlit pandas plotly openpyxl

IMPORTANT: Keep benchmarks.xlsx in the same folder as app.py
"""

import io
import os
import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Energy Benchmarking Platform",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
    .main { background-color: #f8fafc; }
    .block-container { padding-top: 1.5rem; }
    h1 { color: #0f4c81; font-size: 1.6rem !important; }
    h2 { color: #0f4c81; font-size: 1.2rem !important; }
    h3 { font-size: 1rem !important; }
    .flag-pass { background:#f0fdf4; border-left:4px solid #16a34a; padding:10px 14px; border-radius:6px; margin:6px 0; }
    .flag-warn { background:#fffbeb; border-left:4px solid #d97706; padding:10px 14px; border-radius:6px; margin:6px 0; }
    .flag-fail { background:#fef2f2; border-left:4px solid #dc2626; padding:10px 14px; border-radius:6px; margin:6px 0; }
    .bm-card   { background:white; border-radius:10px; padding:16px 20px; border:1px solid #e2e8f0; margin-bottom:12px; }
    .bm-label  { font-size:11px; font-weight:600; color:#64748b; text-transform:uppercase; letter-spacing:.05em; }
    .bm-value  { font-size:22px; font-weight:700; color:#0f4c81; line-height:1.2; }
    .bm-sub    { font-size:12px; color:#94a3b8; margin-top:2px; }
    .success-box { background:#f0fdf4; border:1px solid #bbf7d0; border-radius:8px; padding:12px 16px; margin:8px 0; }
</style>
""", unsafe_allow_html=True)

# ── Constants ─────────────────────────────────────────────────────────────────
# Always look for benchmarks.xlsx in the same folder as this script,
# regardless of which directory the terminal was launched from
BENCHMARK_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "benchmarks.xlsx")
MODEL_TYPES      = ["Proposed","Baseline","Existing"]
PROJECT_PHASES   = ["Concept","Schematic Design","Design Development","100% Design","As-Built"]
SOFTWARE_OPTIONS = ["IES VE","EnergyPlus","OpenStudio","eQUEST","Manual / Excel template","Other"]
DHW_BUILDINGS    = ["School","Office","Hospital","Residential","Community Centre","Library"]
END_USE_COLORS   = ["#ef4444","#3b82f6","#8b5cf6","#f59e0b","#06b6d4","#10b981","#f97316"]

AUTO_MAP = {
    # ── Main energy sources ──
    "electricity_kwh":  ["electricity","total electricity","elec","electricity_kwh","electricity (kwh)",
                         "total elec","elec_kwh","annual electricity","total_kwh","total electricity use"],
    "gas_kwh":          ["gas","natural gas","gas consumption","gas_kwh","naturalgas","natural gas (kwh)",
                         "gas (kwh)","annual gas","ng_kwh","natural_gas","natural_gas_kwh"],
    "area_m2":          ["area","floor area","gfa","area_m2","floor area (m2)","area (m2)",
                         "gross floor area","conditioned area","heated floor area","area_m2"],
    # ── Heating ──
    "heating_kwh":      ["heating","heat","heating_kwh","heating (kwh)","htg","htg energy",
                         "space heating","space_heating","space heating (kwh)","space_heating_kwh",
                         "annual heating","zone heating","heat energy"],
    # ── Cooling ──
    "cooling_kwh":      ["cooling","cool","cooling_kwh","cooling (kwh)","clg","clg energy",
                         "space cooling","space_cooling","space cooling (kwh)","space_cooling_kwh",
                         "annual cooling","zone cooling","heat rejection","heat_rejection",
                         "heat rejection (kwh)","heat_rejection_kwh"],
    # ── Fans — three separate components ──
    "central_fan_kwh":  ["central fan","central_fan","central_fan_kwh","central fan (kwh)",
                         "interior central fan","interior_central_fan","interior_central_fan_kwh",
                         "interior central fans","interior_central_fans","interior_central_fans_kwh",
                         "supply fan","supply_fan","supply_fan_kwh","ahu fan","ahu_fan",
                         "air handling unit","air handling","central air handling",
                         "fans","fan","fans_kwh","fan energy","supply fans","total fan energy"],
    "local_fan_kwh":    ["local fan","local_fan","local_fan_kwh","local fan (kwh)",
                         "interior local fan","interior_local_fan","interior_local_fan_kwh",
                         "interior local fans","interior_local_fans","interior_local_fans_kwh",
                         "fan coil","fan coil unit","fcu","fan_coil","fan_coil_kwh",
                         "zone fan","local ventilation","unit ventilator","local unit"],
    "exhaust_fan_kwh":  ["exhaust fan","exhaust_fan","exhaust_fan_kwh","exhaust fan (kwh)",
                         "exhaust fans","exhaust_fans","exhaust energy","exhaust fan energy",
                         "exhaust_fan_energy","building exhaust","toilet exhaust",
                         "kitchen exhaust","parking exhaust","general exhaust"],
    # ── Lighting ──
    "lighting_kwh":     ["lighting","lights","lighting_kwh","lighting (kwh)","ltg",
                         "interior lighting","interior_lighting","interior_lighting_kwh",
                         "interior lighting (kwh)","annual lighting","zone lighting",
                         "lighting energy","interior lights"],
    # ── DHW ──
    "dhw_kwh":          ["dhw","domestic hot water","dhw_kwh","dhw (kwh)","hot water",
                         "service hot water","shw","dhw energy","dhw_energy",
                         "domestic hot water (kwh)","hot water energy"],
    # ── Pumps ──
    "pumps_kwh":        ["pumps","pump","pumps_kwh","pump energy","pumps (kwh)","heating pumps",
                         "pump_kwh","pumps energy","chw pump","hw pump","condenser pump",
                         "heating pump","cooling pump","pumping energy"],
    # ── Receptacle / plug loads ──
    "receptacle_kwh":   ["receptacle","receptacles","receptacle_kwh","receptacle (kwh)",
                         "plug loads","plug_loads","plug load","plug_load_kwh",
                         "equipment","equipment (kwh)","equipment_kwh","process load",
                         "interior equipment","interior_equipment","interior_equipment_kwh",
                         "miscellaneous","misc load","misc_load","other loads"],
    # ── Interior ceiling / heat rejection misc ──
    "heat_rejection_kwh":["heat rejection","heat_rejection","heat_rejection_kwh",
                          "heat rejection (kwh)","cooling tower","condenser heat",
                          "interior ceiling","interior_ceiling","interior_ce"],
    # ── Exterior lighting ──
    "ext_lighting_kwh": ["exterior lighting","exterior_lighting","exterior_lighting_kwh",
                         "ext lighting","ext_lighting","outdoor lighting","site lighting"],
    # ── Process / other ──
    "process_kwh":      ["process","process load","process_kwh","process (kwh)",
                         "process energy","other","other energy","other_kwh"],
}

FIELD_LABELS = {
    "electricity_kwh":   "Electricity (kWh/yr)",
    "gas_kwh":           "Natural Gas (kWh/yr)",
    "area_m2":           "Floor Area (m²)",
    "heating_kwh":       "Space Heating (kWh/yr)",
    "cooling_kwh":       "Space Cooling (kWh/yr)",
    "central_fan_kwh":   "Interior Central Fan / AHU (kWh/yr)",
    "local_fan_kwh":     "Interior Local Fan (kWh/yr)",
    "exhaust_fan_kwh":   "Exhaust Fan (kWh/yr)",
    "lighting_kwh":      "Interior Lighting (kWh/yr)",
    "dhw_kwh":           "DHW (kWh/yr)",
    "pumps_kwh":         "Pumps (kWh/yr)",
    "receptacle_kwh":    "Receptacle / Plug Loads (kWh/yr)",
    "heat_rejection_kwh":"Heat Rejection (kWh/yr)",
    "ext_lighting_kwh":  "Exterior Lighting (kWh/yr)",
    "process_kwh":       "Process / Other (kWh/yr)",
}

# ── Load benchmarks from Excel ─────────────────────────────────────────────────
def load_benchmarks():
    """Read benchmarks.xlsx and return a dict keyed by (building_type, city, zone)."""
    if not os.path.exists(BENCHMARK_FILE):
        st.error(f"❌ Cannot find **{BENCHMARK_FILE}** — make sure it is in the same folder as app.py")
        st.stop()
    df = pd.read_excel(BENCHMARK_FILE, sheet_name="Benchmarks", dtype=str)
    df.columns = [c.strip() for c in df.columns]
    benchmarks = {}
    for _, row in df.iterrows():
        try:
            btype = str(row["Building Type"]).strip()
            city  = str(row["City"]).strip()
            zone  = str(row["Climate Zone"]).strip()
            pct_raw = str(row["Percentile Data (comma separated)"]).strip()
            pct_data = [float(x.strip()) for x in pct_raw.split(",") if x.strip()]
            benchmarks[(btype, city, zone)] = {
                "good_eui":    float(row["Good EUI"]),
                "median_eui":  float(row["Median EUI"]),
                "high_eui":    float(row["High Flag EUI"]),
                "median_ghgi": float(row["Median GHGI"]),
                "heat_pct":    float(row["Heating %"]),
                "cool_pct":    float(row["Cooling %"]),
                "fan_pct":     float(row["Fan %"]),
                "ltg_pct":     float(row["Lighting %"]),
                "dhw_pct":     float(row["DHW %"]),
                "recept_pct":  float(row["Receptacle %"]),
                "pumps_pct":   float(row["Pumps %"]),
                "pct_data":    pct_data,
            }
        except Exception:
            continue  # skip malformed rows
    return benchmarks

def save_benchmark_to_excel(new_row: dict):
    """Append a new benchmark row to benchmarks.xlsx."""
    wb = load_workbook(BENCHMARK_FILE)
    ws = wb["Benchmarks"]
    next_row = ws.max_row + 1
    values = [
        new_row["building_type"], new_row["city"], new_row["zone"],
        new_row["good_eui"], new_row["median_eui"], new_row["high_eui"],
        new_row["median_ghgi"], new_row["heat_pct"], new_row["cool_pct"],
        new_row["fan_pct"], new_row["ltg_pct"], new_row["dhw_pct"],
        new_row["recept_pct"], new_row["pumps_pct"],
        new_row["pct_data"],
    ]
    row_fill = PatternFill("solid", fgColor="F0F9FF")
    thin = Border(
        left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),  bottom=Side(style="thin", color="D1D5DB"),
    )
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=next_row, column=col_idx, value=val)
        cell.fill   = row_fill
        cell.border = thin
        cell.font   = Font(size=10)
        cell.alignment = Alignment(vertical="center")
    wb.save(BENCHMARK_FILE)

def update_benchmark_in_excel(row_index: int, new_row: dict):
    """Update an existing benchmark row (1-based index into data rows, excluding header)."""
    wb = load_workbook(BENCHMARK_FILE)
    ws = wb["Benchmarks"]
    excel_row = row_index + 1  # +1 for header
    values = [
        new_row["building_type"], new_row["city"], new_row["zone"],
        new_row["good_eui"], new_row["median_eui"], new_row["high_eui"],
        new_row["median_ghgi"], new_row["heat_pct"], new_row["cool_pct"],
        new_row["fan_pct"], new_row["ltg_pct"], new_row["dhw_pct"],
        new_row["recept_pct"], new_row["pumps_pct"], new_row["pct_data"],
    ]
    for col_idx, val in enumerate(values, 1):
        ws.cell(row=excel_row, column=col_idx, value=val)
    wb.save(BENCHMARK_FILE)

def delete_benchmark_in_excel(row_index: int):
    """Delete a benchmark row by its 1-based data index."""
    wb = load_workbook(BENCHMARK_FILE)
    ws = wb["Benchmarks"]
    ws.delete_rows(row_index + 1)
    wb.save(BENCHMARK_FILE)

# ── Helper functions ───────────────────────────────────────────────────────────
def guess_mapping(headers):
    mapping = {}
    for field, aliases in AUTO_MAP.items():
        match = next((h for h in headers if h.strip().lower() in aliases), "")
        mapping[field] = match
    return mapping

def calc_percentile(eui, pct_data):
    above = sum(1 for v in sorted(pct_data) if v > eui)
    return round((above / len(pct_data)) * 100)

def calculate_kpis(vals, area_override=None):
    area    = float(area_override) if area_override else float(vals.get("area_m2") or 1)
    elec    = float(vals.get("electricity_kwh")    or 0)
    gas     = float(vals.get("gas_kwh")            or 0)
    heat    = float(vals.get("heating_kwh")        or 0)
    cool    = float(vals.get("cooling_kwh")        or 0)
    central_fan  = float(vals.get("central_fan_kwh")  or 0)
    local_fan    = float(vals.get("local_fan_kwh")    or 0)
    exhaust_fan  = float(vals.get("exhaust_fan_kwh")  or 0)
    fans         = central_fan + local_fan + exhaust_fan  # total fans = central + local + exhaust
    ltg     = float(vals.get("lighting_kwh")       or 0)
    dhw     = float(vals.get("dhw_kwh")            or 0)
    pumps   = float(vals.get("pumps_kwh")          or 0)
    recept  = float(vals.get("receptacle_kwh")     or 0)
    heat_rej= float(vals.get("heat_rejection_kwh") or 0)
    ext_ltg = float(vals.get("ext_lighting_kwh")   or 0)
    process = float(vals.get("process_kwh")        or 0)
    total   = elec + gas
    def eui(v): return round(v / area, 1) if area else 0
    return {
        "area": area, "total_energy": total,
        "total_eui":      eui(total),
        "elec_eui":       eui(elec),
        "gas_eui":        eui(gas),
        "heat_eui":       eui(heat),
        "cool_eui":       eui(cool),
        "fan_eui":         eui(fans),
        "central_fan_eui": eui(central_fan),
        "local_fan_eui":   eui(local_fan),
        "exhaust_fan_eui": eui(exhaust_fan),
        "ltg_eui":        eui(ltg),
        "dhw_eui":        eui(dhw),
        "pumps_eui":      eui(pumps),
        "recept_eui":     eui(recept),
        "heat_rej_eui":   eui(heat_rej),
        "ext_ltg_eui":    eui(ext_ltg),
        "process_eui":    eui(process),
        "ghgi":           round((elec * 0.00015 + gas * 0.00018) * 1000 / area, 1) if area else 0,
    }

def generate_flags(kpis, bm, building_type):
    flags = []
    if not bm:
        flags.append(("info","ℹ️","No benchmark found for this combination. KPIs calculated but no comparison available."))
        return flags
    fan_bm  = bm["median_eui"] * bm["fan_pct"]  / 100
    heat_bm = bm["median_eui"] * bm["heat_pct"] / 100
    cool_bm = bm["median_eui"] * bm["cool_pct"] / 100
    ltg_bm  = bm["median_eui"] * bm["ltg_pct"]  / 100
    if kpis["total_eui"] > bm["high_eui"]:
        flags.append(("fail","✗",f"Total EUI ({kpis['total_eui']} kWh/m²·yr) exceeds the high flag threshold of {bm['high_eui']} kWh/m²·yr — review all model inputs and fuel assignments."))
    elif kpis["total_eui"] < bm["good_eui"] * 0.55:
        flags.append(("warn","⚠",f"Total EUI ({kpis['total_eui']} kWh/m²·yr) is unusually low — confirm all end-uses are modelled."))
    else:
        flags.append(("pass","✓",f"Total EUI ({kpis['total_eui']} kWh/m²·yr) is within expected range (Good: {bm['good_eui']}, Median: {bm['median_eui']} kWh/m²·yr)."))
    if fan_bm > 0 and kpis["fan_eui"] > fan_bm * 1.35:
        flags.append(("warn","⚠",f"Fan energy ({kpis['fan_eui']} kWh/m²·yr) is {round((kpis['fan_eui']/fan_bm-1)*100)}% above benchmark — verify AHU schedules and fan sizing."))
    elif kpis["fan_eui"] == 0:
        flags.append(("warn","⚠","Fan energy is zero — confirm fan systems are included."))
    if kpis["cool_eui"] == 0:
        flags.append(("warn","⚠","No cooling energy — confirm whether a mechanical cooling system exists."))
    elif cool_bm > 0 and kpis["cool_eui"] < cool_bm * 0.25:
        flags.append(("warn","⚠",f"Cooling energy ({kpis['cool_eui']} kWh/m²·yr) is very low — verify cooling system modelling."))
    if heat_bm > 0 and kpis["heat_eui"] > heat_bm * 1.5:
        flags.append(("warn","⚠",f"Heating energy ({kpis['heat_eui']} kWh/m²·yr) is {round((kpis['heat_eui']/heat_bm-1)*100)}% above benchmark — review envelope and heating schedules."))
    if heat_bm > 0 and cool_bm > 0 and kpis["heat_eui"] > heat_bm * 1.2 and kpis["cool_eui"] > cool_bm * 1.2:
        flags.append(("warn","⚠","Both heating and cooling elevated — possible simultaneous heating/cooling or control issue."))
    if building_type in DHW_BUILDINGS and kpis["dhw_eui"] == 0:
        flags.append(("fail","✗",f"DHW energy is zero for a {building_type} — domestic hot water is typically required."))
    elif kpis["dhw_eui"] > 0:
        flags.append(("pass","✓",f"DHW energy present ({kpis['dhw_eui']} kWh/m²·yr)."))
    if ltg_bm > 0 and kpis["ltg_eui"] > ltg_bm * 1.4:
        flags.append(("warn","⚠",f"Lighting energy ({kpis['ltg_eui']} kWh/m²·yr) above benchmark ({round(ltg_bm,1)}) — verify LPD vs NECB."))
    if kpis["ghgi"] > bm["median_ghgi"] * 1.3:
        flags.append(("warn","⚠",f"GHGI ({kpis['ghgi']} kgCO₂e/m²·yr) above median benchmark ({bm['median_ghgi']}) — review fuel mix."))
    else:
        flags.append(("pass","✓",f"GHGI ({kpis['ghgi']} kgCO₂e/m²·yr) within acceptable range (benchmark median: {bm['median_ghgi']})."))
    return flags

def status_color(val, good, median, high):
    if val <= good:   return "🟢"
    if val <= median: return "🟡"
    if val <= high:   return "🟠"
    return "🔴"

def make_pie(labels, values, title):
    fig = go.Figure(go.Pie(
        labels=labels, values=values, hole=0.42,
        marker=dict(colors=END_USE_COLORS, line=dict(color="#fff", width=2)),
        textinfo="label+percent", textfont=dict(size=12),
        hovertemplate="<b>%{label}</b><br>%{value} kWh/m²·yr<br>%{percent}<extra></extra>",
    ))
    fig.update_layout(
        title=dict(text=title, font=dict(size=13, color="#0f4c81")),
        showlegend=True, legend=dict(orientation="v", x=1.02, y=0.5, font=dict(size=11)),
        margin=dict(t=40,b=10,l=10,r=10), height=320,
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
    )
    return fig

# ── Session state ─────────────────────────────────────────────────────────────
if "step"    not in st.session_state: st.session_state.step    = 1
if "vals"    not in st.session_state: st.session_state.vals    = {}
if "results" not in st.session_state: st.session_state.results = None
if "page"    not in st.session_state: st.session_state.page    = "QA/QC Tool"
if "bm_reload" not in st.session_state: st.session_state.bm_reload = 0

# ── Load benchmarks ───────────────────────────────────────────────────────────
BENCHMARKS = load_benchmarks()
ALL_BUILDING_TYPES = sorted(set(k[0] for k in BENCHMARKS)) or ["School","Office","Retail","Hospital","Residential","Warehouse"]
CITIES             = sorted(set(k[1] for k in BENCHMARKS)) or ["Edmonton","Calgary","Vancouver","Toronto"]
CLIMATE_ZONES      = sorted(set(k[2] for k in BENCHMARKS)) or ["4","5","6","7","8"]

# ── Sidebar ────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⚡ Energy Benchmarking")
    st.markdown("**QA/QC Platform** | MVP v1.0")
    st.divider()
    page = st.radio("Navigate", ["QA/QC Tool", "📚 Benchmark Explorer", "⚙️ Manage Benchmarks"], label_visibility="collapsed")
    st.session_state.page = page
    st.divider()
    if page == "QA/QC Tool":
        steps = ["1. Upload / Enter Data","2. Map Columns","3. Building Info","4. Results"]
        for i, s in enumerate(steps, 1):
            if i < st.session_state.step:    st.markdown(f"✅ {s}")
            elif i == st.session_state.step:  st.markdown(f"**→ {s}**")
            else:                             st.markdown(f"&nbsp;&nbsp;&nbsp;{s}")
        st.divider()
        if st.button("🔄 Start Over", use_container_width=True):
            for k in ["step","vals","results","headers","csv_df","mapping","meta"]:
                if k in st.session_state: del st.session_state[k]
            st.session_state.step = 1
            st.rerun()
    st.divider()
    st.markdown(f"**{len(BENCHMARKS)} benchmarks loaded**")



# ══════════════════════════════════════════════════════════════════════════════
#  PAGE: MANAGE BENCHMARKS
# ══════════════════════════════════════════════════════════════════════════════
if st.session_state.page == "⚙️ Manage Benchmarks":
    st.markdown("# ⚙️ Manage Benchmark Database")
    st.markdown(f"All data is stored in **{BENCHMARK_FILE}** in the same folder as app.py. Changes save immediately.")
    st.divider()

    tab_view, tab_add, tab_edit, tab_delete = st.tabs(["📋 View All", "➕ Add New", "✏️ Edit Existing", "🗑️ Delete"])

    # ── VIEW ──
    with tab_view:
        st.markdown("### All Benchmarks")
        rows = []
        for (btype, city, zone), bm in BENCHMARKS.items():
            rows.append({
                "Building Type": btype, "City": city, "Zone": zone,
                "Good EUI": bm["good_eui"], "Median EUI": bm["median_eui"],
                "High Flag": bm["high_eui"], "GHGI": bm["median_ghgi"],
                "Heating %": bm["heat_pct"], "Cooling %": bm["cool_pct"],
                "Fan %": bm["fan_pct"], "Lighting %": bm["ltg_pct"],
                "DHW %": bm["dhw_pct"], "Receptacle %": bm["recept_pct"],
                "Pumps %": bm["pumps_pct"],
            })
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
        st.caption(f"Total: {len(rows)} benchmark records")

        # Download the Excel file
        with open(BENCHMARK_FILE, "rb") as f:
            st.download_button(
                "📥 Download benchmarks.xlsx",
                data=f.read(),
                file_name="benchmarks.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

    # ── ADD ──
    with tab_add:
        st.markdown("### Add New Benchmark")
        st.info("Fill in all fields below and click **Save to Database**. The new record will be saved to benchmarks.xlsx immediately.")

        c1, c2, c3 = st.columns(3)
        with c1:
            n_btype  = st.text_input("Building Type *", placeholder="e.g. School")
            n_city   = st.text_input("City *",          placeholder="e.g. Edmonton")
            n_zone   = st.text_input("Climate Zone *",  placeholder="e.g. 7")

        with c2:
            n_good   = st.number_input("Good Practice EUI (kWh/m²·yr) *", min_value=0.0, step=5.0)
            n_median = st.number_input("Median EUI (kWh/m²·yr) *",        min_value=0.0, step=5.0)
            n_high   = st.number_input("High Flag EUI (kWh/m²·yr) *",     min_value=0.0, step=5.0)
            n_ghgi   = st.number_input("Median GHGI (kgCO₂e/m²·yr) *",   min_value=0.0, step=1.0)
        with c3:
            st.markdown("**End-Use Percentages** (should sum to ~100%)")
            n_heat   = st.number_input("Heating %",      min_value=0.0, max_value=100.0, step=1.0, value=40.0)
            n_cool   = st.number_input("Cooling %",      min_value=0.0, max_value=100.0, step=1.0, value=8.0)
            n_fan    = st.number_input("Fan %",          min_value=0.0, max_value=100.0, step=1.0, value=15.0)
            n_ltg    = st.number_input("Lighting %",     min_value=0.0, max_value=100.0, step=1.0, value=22.0)
            n_dhw    = st.number_input("DHW %",          min_value=0.0, max_value=100.0, step=1.0, value=8.0)
            n_recept = st.number_input("Receptacle %",   min_value=0.0, max_value=100.0, step=1.0, value=5.0)
            n_pumps  = st.number_input("Pumps %",        min_value=0.0, max_value=100.0, step=1.0, value=2.0)
            pct_sum = n_heat + n_cool + n_fan + n_ltg + n_dhw + n_recept + n_pumps
            color = "🟢" if 95 <= pct_sum <= 105 else "🔴"
            st.markdown(f"{color} Percentages sum to **{pct_sum:.0f}%** (target: ~100%)")

        st.markdown("**Percentile Data** — 10 EUI values representing 10th to 100th percentile, comma separated")
        n_pct_raw = st.text_input("Percentile Data *", placeholder="e.g. 95,110,125,141,155,165,180,195,210,230")

        if st.button("💾 Save to Database", type="primary", use_container_width=True):
            errors = []
            if not n_btype.strip(): errors.append("Building Type is required")
            if not n_city.strip():  errors.append("City is required")
            if not n_zone.strip():  errors.append("Climate Zone is required")
            if n_median <= 0:       errors.append("Median EUI must be greater than 0")
            if not n_pct_raw.strip(): errors.append("Percentile Data is required")
            else:
                try:
                    pct_vals = [float(x.strip()) for x in n_pct_raw.split(",") if x.strip()]
                    if len(pct_vals) != 10: errors.append("Percentile Data must have exactly 10 values")
                except:
                    errors.append("Percentile Data must be numbers separated by commas")

            # Check for duplicate
            key = (n_btype.strip(), n_city.strip(), n_zone.strip())
            if key in BENCHMARKS: errors.append(f"A benchmark for {n_btype} · {n_city} · Zone {n_zone} already exists. Use Edit to update it.")

            if errors:
                for e in errors: st.error(e)
            else:
                save_benchmark_to_excel({
                    "building_type": n_btype.strip(), "city": n_city.strip(), "zone": n_zone.strip(),
                    "good_eui": n_good, "median_eui": n_median, "high_eui": n_high, "median_ghgi": n_ghgi,
                    "heat_pct": n_heat, "cool_pct": n_cool, "fan_pct": n_fan,
                    "ltg_pct": n_ltg, "dhw_pct": n_dhw, "recept_pct": n_recept, "pumps_pct": n_pumps,
                    "pct_data": n_pct_raw.strip(),
                })
                st.success(f"✅ Benchmark added: **{n_btype} · {n_city} · Zone {n_zone}**. Reloading...")
                st.session_state.bm_reload += 1
                st.rerun()

    # ── EDIT ──
    with tab_edit:
        st.markdown("### Edit Existing Benchmark")
        bm_keys   = list(BENCHMARKS.keys())
        bm_labels = [f"{k[0]} · {k[1]} · Zone {k[2]}" for k in bm_keys]
        selected  = st.selectbox("Select benchmark to edit", bm_labels)
        sel_idx   = bm_labels.index(selected)
        sel_key   = bm_keys[sel_idx]
        bm        = BENCHMARKS[sel_key]

        st.markdown(f"Editing: **{selected}**")
        ec1, ec2, ec3 = st.columns(3)
        with ec1:
            e_btype  = st.text_input("Building Type",  value=sel_key[0], key="e_btype")
            e_city   = st.text_input("City",           value=sel_key[1], key="e_city")
            e_zone   = st.text_input("Climate Zone",   value=sel_key[2], key="e_zone")

        with ec2:
            e_good   = st.number_input("Good EUI",   value=bm["good_eui"],   step=5.0, key="e_good")
            e_median = st.number_input("Median EUI", value=bm["median_eui"], step=5.0, key="e_median")
            e_high   = st.number_input("High Flag",  value=bm["high_eui"],   step=5.0, key="e_high")
            e_ghgi   = st.number_input("Median GHGI",value=bm["median_ghgi"],step=1.0, key="e_ghgi")
        with ec3:
            st.markdown("**End-Use Percentages**")
            e_heat   = st.number_input("Heating %",    value=bm["heat_pct"],   step=1.0, key="e_heat")
            e_cool   = st.number_input("Cooling %",    value=bm["cool_pct"],   step=1.0, key="e_cool")
            e_fan    = st.number_input("Fan %",        value=bm["fan_pct"],    step=1.0, key="e_fan")
            e_ltg    = st.number_input("Lighting %",   value=bm["ltg_pct"],    step=1.0, key="e_ltg")
            e_dhw    = st.number_input("DHW %",        value=bm["dhw_pct"],    step=1.0, key="e_dhw")
            e_recept = st.number_input("Receptacle %", value=bm["recept_pct"], step=1.0, key="e_recept")
            e_pumps  = st.number_input("Pumps %",      value=bm["pumps_pct"],  step=1.0, key="e_pumps")
            e_sum    = e_heat + e_cool + e_fan + e_ltg + e_dhw + e_recept + e_pumps
            st.markdown(f"{'🟢' if 95<=e_sum<=105 else '🔴'} Sum: **{e_sum:.0f}%**")

        e_pct_raw = st.text_input("Percentile Data (10 values, comma separated)", value=",".join(str(int(v)) for v in bm["pct_data"]), key="e_pct")

        if st.button("💾 Save Changes", type="primary", use_container_width=True):
            update_benchmark_in_excel(sel_idx + 1, {
                "building_type": e_btype, "city": e_city, "zone": e_zone,
                "good_eui": e_good, "median_eui": e_median, "high_eui": e_high, "median_ghgi": e_ghgi,
                "heat_pct": e_heat, "cool_pct": e_cool, "fan_pct": e_fan,
                "ltg_pct": e_ltg, "dhw_pct": e_dhw, "recept_pct": e_recept, "pumps_pct": e_pumps,
                "pct_data": e_pct_raw,
            })
            st.success(f"✅ Benchmark updated: **{e_btype} · {e_city} · Zone {e_zone}**. Reloading...")
            st.rerun()

    # ── DELETE ──
    with tab_delete:
        st.markdown("### Delete a Benchmark")
        st.warning("⚠️ This permanently removes the record from benchmarks.xlsx. This cannot be undone.")
        bm_keys   = list(BENCHMARKS.keys())
        bm_labels = [f"{k[0]} · {k[1]} · Zone {k[2]}" for k in bm_keys]
        del_sel   = st.selectbox("Select benchmark to delete", bm_labels, key="del_sel")
        del_idx   = bm_labels.index(del_sel)

        col_confirm, col_btn = st.columns([3,1])
        with col_confirm:
            confirm = st.checkbox(f"I confirm I want to permanently delete **{del_sel}**")
        with col_btn:
            if st.button("🗑️ Delete", type="primary", disabled=not confirm, use_container_width=True):
                delete_benchmark_in_excel(del_idx + 1)
                st.success(f"✅ Deleted: **{del_sel}**. Reloading...")
                st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
#  PAGE: BENCHMARK EXPLORER
# ══════════════════════════════════════════════════════════════════════════════
elif st.session_state.page == "📚 Benchmark Explorer":
    st.markdown("# 📚 Benchmark Explorer")
    st.markdown("Browse the benchmark database by building type and location — no upload required.")
    st.divider()

    cf1, cf2, cf3 = st.columns(3)
    with cf1: bx_type = st.selectbox("Building Type", ALL_BUILDING_TYPES)
    with cf2: bx_city = st.selectbox("City", CITIES + ["All cities"])
    with cf3: bx_zone = st.selectbox("Climate Zone", CLIMATE_ZONES, index=min(3, len(CLIMATE_ZONES)-1))

    matches = {k:v for k,v in BENCHMARKS.items() if k[0]==bx_type and (bx_city=="All cities" or k[1]==bx_city) and k[2]==bx_zone}

    if not matches:
        st.warning(f"No benchmark data found for **{bx_type}** in **{bx_city}** (Zone {bx_zone}). Try a different combination, or add this benchmark in ⚙️ Manage Benchmarks.")

        st.stop()

    for (btype, bcity, bzone), bm in matches.items():
        st.markdown(f"## 🏢 {btype} · {bcity} · Climate Zone {bzone}")
        ci1,ci2 = st.columns(2)
        ci1.markdown(f'<div class="bm-card"><div class="bm-label">Median GHGI</div><div class="bm-value">{bm["median_ghgi"]}</div><div class="bm-sub">kgCO₂e/m²·yr</div></div>', unsafe_allow_html=True)
        ci2.markdown(f'<div class="bm-card"><div class="bm-label">Total Benchmark Records</div><div class="bm-value" style="font-size:15px">{len(BENCHMARKS)}</div><div class="bm-sub">in database</div></div>', unsafe_allow_html=True)

        m1,m2,m3 = st.columns(3)
        m1.metric("🟢 Good Practice EUI", f"{bm['good_eui']} kWh/m²·yr")
        m2.metric("🟡 Median EUI",        f"{bm['median_eui']} kWh/m²·yr")
        m3.metric("🔴 High Flag EUI",     f"{bm['high_eui']} kWh/m²·yr")

        eu_labels = ["Heating","Cooling","Fans","Lighting","DHW","Receptacle","Pumps"]
        eu_pcts   = [bm["heat_pct"],bm["cool_pct"],bm["fan_pct"],bm["ltg_pct"],bm["dhw_pct"],bm["recept_pct"],bm["pumps_pct"]]
        med_vals  = [round(bm["median_eui"]*p/100,1) for p in eu_pcts]
        good_vals = [round(bm["good_eui"]*p/100,1)   for p in eu_pcts]
        high_vals = [round(bm["high_eui"]*p/100,1)   for p in eu_pcts]

        cp1, cp2 = st.columns(2)
        with cp1:
            pie_l = [l for l,v in zip(eu_labels,med_vals) if v>0]
            pie_v = [v for v in med_vals if v>0]
            st.plotly_chart(make_pie(pie_l, pie_v, f"Median End-Use Split — {btype} · {bcity}"), use_container_width=True)
        with cp2:
            fig_bar = go.Figure()
            fig_bar.add_trace(go.Bar(name="Good Practice", x=eu_labels, y=good_vals, marker_color="#16a34a", opacity=0.85))
            fig_bar.add_trace(go.Bar(name="Median",        x=eu_labels, y=med_vals,  marker_color="#d97706", opacity=0.85))
            fig_bar.add_trace(go.Bar(name="High Flag",     x=eu_labels, y=high_vals, marker_color="#dc2626", opacity=0.85))
            fig_bar.update_layout(barmode="group", template="plotly_white", height=320,
                                  title=dict(text="End-Use EUI by Performance Tier", font=dict(size=13,color="#0f4c81")),
                                  yaxis_title="EUI (kWh/m²·yr)", legend=dict(orientation="h",y=-0.25), margin=dict(t=40,b=10))
            st.plotly_chart(fig_bar, use_container_width=True)

        st.markdown("#### End-Use Breakdown Table")
        st.dataframe(pd.DataFrame({"End Use":eu_labels,"Share (%)":eu_pcts,
            "Good Practice (kWh/m²·yr)":good_vals,"Median (kWh/m²·yr)":med_vals,"High Flag (kWh/m²·yr)":high_vals}),
            use_container_width=True, hide_index=True)

        st.markdown("#### Portfolio Percentile Distribution")
        sorted_pcts = sorted(bm["pct_data"])
        bar_colors  = ["#16a34a" if v<=bm["good_eui"] else "#d97706" if v<=bm["median_eui"] else "#ea580c" if v<=bm["high_eui"] else "#dc2626" for v in sorted_pcts]
        fig_pct = go.Figure(go.Bar(x=[f"{i*10}th" for i in range(1,11)], y=sorted_pcts, marker_color=bar_colors,
            hovertemplate="<b>%{x} percentile</b><br>EUI: %{y} kWh/m²·yr<extra></extra>"))
        fig_pct.add_hline(y=bm["good_eui"],   line_dash="dot",  line_color="#16a34a", line_width=1.5, annotation_text=f"  Good: {bm['good_eui']}",   annotation_font_color="#16a34a")
        fig_pct.add_hline(y=bm["median_eui"], line_dash="dash", line_color="#d97706", line_width=1.5, annotation_text=f"  Median: {bm['median_eui']}", annotation_font_color="#d97706")
        fig_pct.add_hline(y=bm["high_eui"],   line_dash="dot",  line_color="#dc2626", line_width=1.5, annotation_text=f"  High: {bm['high_eui']}",     annotation_font_color="#dc2626")
        fig_pct.update_layout(template="plotly_white", height=300, xaxis_title="Percentile", yaxis_title="EUI (kWh/m²·yr)", margin=dict(t=20,b=10))
        st.plotly_chart(fig_pct, use_container_width=True)

        st.divider()


# ══════════════════════════════════════════════════════════════════════════════
#  PAGE: QA/QC TOOL
# ══════════════════════════════════════════════════════════════════════════════
else:
    st.markdown("# ⚡ Energy Model Benchmarking & QA/QC Platform")
    st.markdown("Upload simulation results → Calculate KPIs → Compare to benchmarks → Get QA/QC flags")
    st.divider()

    if st.session_state.step == 1:
        st.markdown("## Step 1 — Upload Simulation Results or Enter Manually")
        tab_upload, tab_manual, tab_sample = st.tabs(["📂 Upload CSV","⌨️ Enter Manually","📋 Load Sample Data"])
        with tab_upload:
            uploaded = st.file_uploader("Choose CSV file", type=["csv"], label_visibility="collapsed")
            if uploaded:
                df = pd.read_csv(uploaded)
                st.session_state.csv_df  = df
                st.session_state.headers = list(df.columns)
                st.success(f"✅ {uploaded.name} — {len(df)} rows, {len(df.columns)} columns")
                st.dataframe(df.head(3), use_container_width=True)
                if st.button("Next: Map Columns →", type="primary"):
                    st.session_state.step = 2; st.rerun()
        with tab_manual:
            c1,c2,c3 = st.columns(3)
            manual = {}
            with c1:
                st.markdown("**Energy Sources**")
                manual["electricity_kwh"]   = st.number_input("Electricity (kWh/yr)",        min_value=0.0, step=1000.0, format="%.0f")
                manual["gas_kwh"]           = st.number_input("Natural Gas (kWh/yr)",        min_value=0.0, step=1000.0, format="%.0f")
                manual["area_m2"]           = st.number_input("Floor Area (m²)",              min_value=0.0, step=100.0,  format="%.0f")
            with c2:
                st.markdown("**HVAC End Uses**")
                manual["heating_kwh"]       = st.number_input("Space Heating (kWh/yr)",      min_value=0.0, step=1000.0, format="%.0f")
                manual["cooling_kwh"]       = st.number_input("Space Cooling (kWh/yr)",      min_value=0.0, step=1000.0, format="%.0f")
                manual["central_fan_kwh"]  = st.number_input("Interior Central Fan / AHU (kWh/yr)", min_value=0.0, step=1000.0, format="%.0f")
                manual["local_fan_kwh"]    = st.number_input("Interior Local Fan (kWh/yr)",        min_value=0.0, step=1000.0, format="%.0f")
                manual["exhaust_fan_kwh"]  = st.number_input("Exhaust Fan (kWh/yr)",               min_value=0.0, step=1000.0, format="%.0f")
                manual["pumps_kwh"]         = st.number_input("Pumps (kWh/yr)",              min_value=0.0, step=500.0,  format="%.0f")
                manual["heat_rejection_kwh"]= st.number_input("Heat Rejection (kWh/yr)",     min_value=0.0, step=500.0,  format="%.0f")
            with c3:
                st.markdown("**Other End Uses**")
                manual["lighting_kwh"]      = st.number_input("Interior Lighting (kWh/yr)",  min_value=0.0, step=1000.0, format="%.0f")
                manual["dhw_kwh"]           = st.number_input("DHW (kWh/yr)",                min_value=0.0, step=500.0,  format="%.0f")
                manual["receptacle_kwh"]    = st.number_input("Receptacle / Plug Loads (kWh/yr)", min_value=0.0, step=500.0, format="%.0f")
                manual["ext_lighting_kwh"]  = st.number_input("Exterior Lighting (kWh/yr)",  min_value=0.0, step=500.0,  format="%.0f")
                manual["process_kwh"]       = st.number_input("Process / Other (kWh/yr)",    min_value=0.0, step=500.0,  format="%.0f")
            if st.button("Next: Building Info →", type="primary"):
                st.session_state.vals = manual; st.session_state.step = 3; st.rerun()
        with tab_sample:
            st.markdown("**Edmonton school, 8,500 m², Boiler + VAV** — DHW is zero to trigger a QA flag example")
            if st.button("Load Sample & Continue →", type="primary"):
                st.session_state.vals = {"electricity_kwh":720000,"gas_kwh":480000,"area_m2":8500,
                    "heating_kwh":380000,"cooling_kwh":25500,
                    "central_fan_kwh":170000,"local_fan_kwh":45000,"exhaust_fan_kwh":25000,
                    "lighting_kwh":170000,"dhw_kwh":0,"pumps_kwh":42000}
                st.session_state.step = 3; st.rerun()

    elif st.session_state.step == 2:
        st.markdown("## Step 2 — Map Columns")
        headers = st.session_state.headers
        auto_map = guess_mapping(headers)
        options  = ["— not mapped —"] + headers
        mapping  = {}
        auto_matched = sum(1 for v in auto_map.values() if v)
        st.info(f"Auto-matched {auto_matched} of {len(FIELD_LABELS)} fields. Review and fix any that say '— not mapped —'.")

        # Group fields into sections for clarity
        sections = {
            "⚡ Energy Sources": ["electricity_kwh","gas_kwh","area_m2"],
            "🔥 HVAC End Uses":  ["heating_kwh","cooling_kwh","central_fan_kwh","local_fan_kwh","exhaust_fan_kwh","pumps_kwh","heat_rejection_kwh"],
            "💡 Other End Uses": ["lighting_kwh","dhw_kwh","receptacle_kwh","ext_lighting_kwh","process_kwh"],
        }
        for section_title, keys in sections.items():
            st.markdown(f"**{section_title}**")
            cols = st.columns(3)
            for i, key in enumerate(keys):
                label = FIELD_LABELS.get(key, key)
                with cols[i % 3]:
                    default_idx = options.index(auto_map[key]) if auto_map.get(key) and auto_map[key] in options else 0
                    sel = st.selectbox(label, options, index=default_idx, key=f"map_{key}")
                    mapping[key] = sel if sel != "— not mapped —" else None
            st.markdown("")
        cb,cn = st.columns([1,4])
        with cb:
            if st.button("← Back"): st.session_state.step=1; st.rerun()
        with cn:
            if st.button("Next: Building Info →", type="primary"):
                df = st.session_state.csv_df
                vals = {key: pd.to_numeric(df[col], errors="coerce").sum() if col else 0 for key,col in mapping.items()}
                st.session_state.vals=vals; st.session_state.mapping=mapping; st.session_state.step=3; st.rerun()

    elif st.session_state.step == 3:
        st.markdown("## Step 3 — Building Information")
        c1,c2 = st.columns(2)
        with c1:
            project_name  = st.text_input("Project Name", placeholder="e.g. New School A")
            building_type = st.selectbox("Building Type", ALL_BUILDING_TYPES)
            city          = st.selectbox("City", CITIES)
            climate_zone  = st.selectbox("Climate Zone", CLIMATE_ZONES, index=min(3,len(CLIMATE_ZONES)-1))
        with c2:
            software      = st.selectbox("Simulation Software", SOFTWARE_OPTIONS)
            model_type    = st.selectbox("Model Type", MODEL_TYPES)
            phase         = st.selectbox("Project Phase", PROJECT_PHASES, index=3)
            area_override = st.number_input("Floor Area Override (m²) — 0 = use mapped value", min_value=0.0, step=100.0, format="%.0f")
        cb,cn = st.columns([1,4])
        with cb:
            if st.button("← Back"):
                st.session_state.step = 2 if "headers" in st.session_state else 1; st.rerun()
        with cn:
            if st.button("Calculate KPIs & View Results →", type="primary"):
                kpis   = calculate_kpis(st.session_state.vals, area_override if area_override>0 else None)
                bm_key = (building_type, city, climate_zone)
                bm     = BENCHMARKS.get(bm_key)
                flags  = generate_flags(kpis, bm, building_type)
                pct    = calc_percentile(kpis["total_eui"], bm["pct_data"]) if bm else None
                st.session_state.results = {"kpis":kpis,"bm":bm,"flags":flags,"percentile":pct,
                    "meta":{"project_name":project_name,"building_type":building_type,"city":city,
                            "climate_zone":climate_zone,"software":software,"model_type":model_type,
                            "phase":phase,"date":str(date.today())}}
                st.session_state.step=4; st.rerun()

    elif st.session_state.step == 4 and st.session_state.results:
        r=st.session_state.results; kpis=r["kpis"]; bm=r["bm"]; meta=r["meta"]; flags=r["flags"]; pct=r["percentile"]

        # ── Project header banner ──
        fc={"pass":0,"warn":0,"fail":0}
        for f in flags: fc[f[0]]=fc.get(f[0],0)+1
        overall       = "Issues Found"    if fc["fail"]>0 else "Review Required" if fc["warn"]>0 else "All Clear"
        overall_icon  = "🔴"              if fc["fail"]>0 else "🟡"              if fc["warn"]>0 else "🟢"
        overall_color = "#fef2f2"         if fc["fail"]>0 else "#fffbeb"         if fc["warn"]>0 else "#f0fdf4"
        overall_border= "#dc2626"         if fc["fail"]>0 else "#d97706"         if fc["warn"]>0 else "#16a34a"
        proj_name     = meta["project_name"] or "Project Results"
        pct_text      = f"&nbsp;&nbsp;|&nbsp;&nbsp; Benchmark percentile: <b>{pct}th</b> (lower = better)" if pct else ""
        flag_text     = f"Pass: {fc['pass']}  &nbsp; Review: {fc['warn']}  &nbsp; Fail: {fc['fail']}"
        info_line     = (f"{meta['building_type']} &middot; {meta['city']} &middot; "
                         f"Climate Zone {meta['climate_zone']} &middot; {meta['model_type']} "
                         f"&middot; {meta['phase']} &middot; {meta['software']}")
        banner_html = (
            f'<div style="background:{overall_color};border-left:5px solid {overall_border};'
            f'border-radius:10px;padding:18px 22px;margin-bottom:16px">'
            f'<div style="font-size:22px;font-weight:700;color:#1e293b">{proj_name}</div>'
            f'<div style="font-size:13px;color:#64748b;margin-top:4px">{info_line}</div>'
            f'<div style="margin-top:12px;display:flex;gap:8px;align-items:center;flex-wrap:wrap">'
            f'<span style="font-size:14px;font-weight:700;color:{overall_border}">{overall_icon} {overall}</span>'
            f'<span style="font-size:13px;color:#64748b">{pct_text}</span>'
            f'</div>'
            f'<div style="margin-top:6px;font-size:13px;color:#94a3b8">{flag_text}</div>'
            f'</div>'
        )
        st.markdown(banner_html, unsafe_allow_html=True)

        # ── Section 1: Energy Summary ──
        st.markdown("### 📊 Energy Summary")
        st.caption("How much energy does this building use per square metre per year?")

        # Row 1 — main totals
        r1c1,r1c2,r1c3,r1c4 = st.columns(4)
        with r1c1:
            delta_eui = f"Benchmark median: {bm['median_eui']} kWh/m²·yr" if bm else "No benchmark"
            color_eui = "#f0fdf4" if bm and kpis["total_eui"]<=bm["good_eui"] else "#fef2f2" if bm and kpis["total_eui"]>bm["high_eui"] else "#fffbeb"
            st.markdown(f'''<div style="background:{color_eui};border-radius:10px;padding:14px 16px;border:1px solid #e2e8f0">
                <div style="font-size:11px;font-weight:600;color:#64748b;text-transform:uppercase;letter-spacing:.05em">Total EUI</div>
                <div style="font-size:28px;font-weight:700;color:#0f4c81;line-height:1.1">{kpis["total_eui"]}</div>
                <div style="font-size:12px;color:#64748b">kWh/m²·yr</div>
                <div style="font-size:12px;color:#94a3b8;margin-top:4px">{delta_eui}</div>
            </div>''', unsafe_allow_html=True)
        with r1c2:
            st.markdown(f'''<div style="background:#f8fafc;border-radius:10px;padding:14px 16px;border:1px solid #e2e8f0">
                <div style="font-size:11px;font-weight:600;color:#64748b;text-transform:uppercase;letter-spacing:.05em">GHGI</div>
                <div style="font-size:28px;font-weight:700;color:#0f4c81;line-height:1.1">{kpis["ghgi"]}</div>
                <div style="font-size:12px;color:#64748b">kgCO₂e / m²·yr</div>
                <div style="font-size:12px;color:#94a3b8;margin-top:4px">{"Benchmark median: " + str(bm["median_ghgi"]) + " kgCO₂e/m²·yr" if bm else "No benchmark"}</div>
            </div>''', unsafe_allow_html=True)
        with r1c3:
            st.markdown(f'''<div style="background:#f8fafc;border-radius:10px;padding:14px 16px;border:1px solid #e2e8f0">
                <div style="font-size:11px;font-weight:600;color:#64748b;text-transform:uppercase;letter-spacing:.05em">Electricity EUI</div>
                <div style="font-size:28px;font-weight:700;color:#3b82f6;line-height:1.1">{kpis["elec_eui"]}</div>
                <div style="font-size:12px;color:#64748b">kWh/m²·yr</div>
                <div style="font-size:12px;color:#94a3b8;margin-top:4px">Floor area: {round(kpis["area"]):,} m²</div>
            </div>''', unsafe_allow_html=True)
        with r1c4:
            st.markdown(f'''<div style="background:#f8fafc;border-radius:10px;padding:14px 16px;border:1px solid #e2e8f0">
                <div style="font-size:11px;font-weight:600;color:#64748b;text-transform:uppercase;letter-spacing:.05em">Gas EUI</div>
                <div style="font-size:28px;font-weight:700;color:#ef4444;line-height:1.1">{kpis["gas_eui"]}</div>
                <div style="font-size:12px;color:#64748b">kWh/m²·yr</div>
                <div style="font-size:12px;color:#94a3b8;margin-top:4px">Total energy: {round(kpis["total_energy"]/1000):,} MWh/yr</div>
            </div>''', unsafe_allow_html=True)

        st.markdown("")

        # Row 2 — end-use breakdown
        r2c1,r2c2,r2c3,r2c4,r2c5 = st.columns(5)
        end_use_cards = [
            ("Heating EUI",      kpis["heat_eui"],                "#ef4444", bm["median_eui"]*bm["heat_pct"]/100   if bm else None),
            ("Cooling EUI",      kpis["cool_eui"],                "#3b82f6", bm["median_eui"]*bm["cool_pct"]/100   if bm else None),
            ("Central Fan EUI",  kpis.get("central_fan_eui",0),  "#8b5cf6", bm["median_eui"]*bm["fan_pct"]/100/3 if bm else None),
            ("Local Fan EUI",    kpis.get("local_fan_eui",0),    "#a78bfa", bm["median_eui"]*bm["fan_pct"]/100/3 if bm else None),
            ("Exhaust Fan EUI",  kpis.get("exhaust_fan_eui",0),  "#7c3aed", bm["median_eui"]*bm["fan_pct"]/100/3 if bm else None),
        ]
        for col, (label, val, color, bm_val) in zip([r2c1,r2c2,r2c3,r2c4,r2c5], end_use_cards):
            bm_text = f"Benchmark: {round(bm_val,1)}" if bm_val is not None else ""
            col.markdown(f'''<div style="background:#f8fafc;border-radius:10px;padding:12px 14px;border:1px solid #e2e8f0;border-top:3px solid {color}">
                <div style="font-size:11px;font-weight:600;color:#64748b;text-transform:uppercase;letter-spacing:.05em">{label}</div>
                <div style="font-size:22px;font-weight:700;color:{color};line-height:1.1">{val}</div>
                <div style="font-size:11px;color:#94a3b8;margin-top:2px">kWh/m²·yr</div>
                <div style="font-size:11px;color:#94a3b8">{bm_text}</div>
            </div>''', unsafe_allow_html=True)

        st.divider()

        # ── Section 2: Charts ──
        st.markdown("### 📈 Charts")
        cc1,cc2 = st.columns(2)
        with cc1:
            eu_l=["Heating","Cooling","Fans","Lighting","DHW","Receptacle","Pumps"]
            eu_v=[kpis["heat_eui"],kpis["cool_eui"],kpis["fan_eui"],kpis["ltg_eui"],kpis["dhw_eui"],kpis.get("recept_eui",0),kpis["pumps_eui"]]
            pl=[l for l,v in zip(eu_l,eu_v) if v>0]; pv=[v for v in eu_v if v>0]
            st.plotly_chart(make_pie(pl,pv,"End-Use Energy Split"), use_container_width=True)
        with cc2:
            if bm:
                sorted_pcts=sorted(bm["pct_data"])
                bar_colors=["#16a34a" if v<=bm["good_eui"] else "#d97706" if v<=bm["median_eui"] else "#ea580c" if v<=bm["high_eui"] else "#dc2626" for v in sorted_pcts]
                fig_pct=go.Figure(go.Bar(x=[f"{i*10}th" for i in range(1,11)],y=sorted_pcts,marker_color=bar_colors,
                    hovertemplate="<b>%{x} percentile</b><br>EUI: %{y} kWh/m²·yr<extra></extra>"))
                fig_pct.add_hline(y=kpis["total_eui"],line_dash="dash",line_color="#0f4c81",line_width=2.5,
                    annotation_text=f"  Your model: {kpis['total_eui']}",annotation_font=dict(color="#0f4c81",size=12))
                fig_pct.update_layout(template="plotly_white",height=320,
                    title=dict(text="Where does your building rank? (lower = better)",font=dict(size=13,color="#0f4c81")),
                    xaxis_title="Percentile rank of similar buildings",yaxis_title="EUI (kWh/m²·yr)",margin=dict(t=45,b=10))
                st.plotly_chart(fig_pct,use_container_width=True)

        if bm:
            eu_l2=["Heating","Cooling","Central Fan","Local Fan","Exhaust Fan","Lighting","DHW","Receptacle","Pumps"]
            your_eu=[kpis["heat_eui"],kpis["cool_eui"],kpis.get("central_fan_eui",0),kpis.get("local_fan_eui",0),
                     kpis.get("exhaust_fan_eui",0),kpis["ltg_eui"],kpis["dhw_eui"],kpis.get("recept_eui",0),kpis["pumps_eui"]]
            fan_split = bm["fan_pct"] / 3  # split benchmark fan % evenly across central/local/exhaust
            bm_good=[round(bm["good_eui"]*p/100,1)  for p in [bm["heat_pct"],bm["cool_pct"],fan_split,fan_split,fan_split,bm["ltg_pct"],bm["dhw_pct"],bm["recept_pct"],bm["pumps_pct"]]]
            bm_med =[round(bm["median_eui"]*p/100,1) for p in [bm["heat_pct"],bm["cool_pct"],fan_split,fan_split,fan_split,bm["ltg_pct"],bm["dhw_pct"],bm["recept_pct"],bm["pumps_pct"]]]
            fig_cmp=go.Figure()
            fig_cmp.add_trace(go.Bar(name="Your Model",    x=eu_l2,y=your_eu, marker_color="#0f4c81"))
            fig_cmp.add_trace(go.Bar(name="Good Practice", x=eu_l2,y=bm_good, marker_color="#16a34a",opacity=0.75))
            fig_cmp.add_trace(go.Bar(name="Median",        x=eu_l2,y=bm_med,  marker_color="#d97706",opacity=0.75))
            fig_cmp.update_layout(barmode="group",template="plotly_white",height=320,
                title=dict(text="Your End-Uses vs Benchmark",font=dict(size=13,color="#0f4c81")),
                yaxis_title="EUI (kWh/m²·yr)",legend=dict(orientation="h",y=-0.25),margin=dict(t=40,b=10))
            st.plotly_chart(fig_cmp,use_container_width=True)

        st.divider()

        # ── Section 3: Benchmark Table ──
        if bm:
            st.markdown("### 📋 Detailed Benchmark Comparison")
            st.caption("🟢 Good practice &nbsp; 🟡 Typical range &nbsp; 🟠 Above median &nbsp; 🔴 Above high flag")
            rows=[
                {"Metric":"Total EUI (kWh/m²·yr)",     "Your Model":kpis["total_eui"],        "Good Practice":bm["good_eui"],                             "Median":bm["median_eui"],                             "High Flag":bm["high_eui"],                             "Status":status_color(kpis["total_eui"],bm["good_eui"],bm["median_eui"],bm["high_eui"])},
                {"Metric":"Heating EUI (kWh/m²·yr)",   "Your Model":kpis["heat_eui"],         "Good Practice":round(bm["good_eui"]*bm["heat_pct"]/100),   "Median":round(bm["median_eui"]*bm["heat_pct"]/100),   "High Flag":round(bm["high_eui"]*bm["heat_pct"]/100),   "Status":status_color(kpis["heat_eui"],  round(bm["good_eui"]*bm["heat_pct"]/100),  round(bm["median_eui"]*bm["heat_pct"]/100),  round(bm["high_eui"]*bm["heat_pct"]/100))},
                {"Metric":"Cooling EUI (kWh/m²·yr)",   "Your Model":kpis["cool_eui"],         "Good Practice":round(bm["good_eui"]*bm["cool_pct"]/100),   "Median":round(bm["median_eui"]*bm["cool_pct"]/100),   "High Flag":round(bm["high_eui"]*bm["cool_pct"]/100),   "Status":status_color(kpis["cool_eui"],  round(bm["good_eui"]*bm["cool_pct"]/100),  round(bm["median_eui"]*bm["cool_pct"]/100),  round(bm["high_eui"]*bm["cool_pct"]/100))},
                {"Metric":"Fan EUI (kWh/m²·yr)",       "Your Model":kpis["fan_eui"],          "Good Practice":round(bm["good_eui"]*bm["fan_pct"]/100),    "Median":round(bm["median_eui"]*bm["fan_pct"]/100),    "High Flag":round(bm["high_eui"]*bm["fan_pct"]/100),    "Status":status_color(kpis["fan_eui"],   round(bm["good_eui"]*bm["fan_pct"]/100),   round(bm["median_eui"]*bm["fan_pct"]/100),   round(bm["high_eui"]*bm["fan_pct"]/100))},
                {"Metric":"Lighting EUI (kWh/m²·yr)",  "Your Model":kpis["ltg_eui"],          "Good Practice":round(bm["good_eui"]*bm["ltg_pct"]/100),    "Median":round(bm["median_eui"]*bm["ltg_pct"]/100),    "High Flag":round(bm["high_eui"]*bm["ltg_pct"]/100),    "Status":status_color(kpis["ltg_eui"],   round(bm["good_eui"]*bm["ltg_pct"]/100),   round(bm["median_eui"]*bm["ltg_pct"]/100),   round(bm["high_eui"]*bm["ltg_pct"]/100))},
                {"Metric":"DHW EUI (kWh/m²·yr)",       "Your Model":kpis["dhw_eui"],          "Good Practice":round(bm["good_eui"]*bm["dhw_pct"]/100),    "Median":round(bm["median_eui"]*bm["dhw_pct"]/100),    "High Flag":round(bm["high_eui"]*bm["dhw_pct"]/100),    "Status":status_color(kpis["dhw_eui"],   round(bm["good_eui"]*bm["dhw_pct"]/100),   round(bm["median_eui"]*bm["dhw_pct"]/100),   round(bm["high_eui"]*bm["dhw_pct"]/100))},
                {"Metric":"Receptacle EUI (kWh/m²·yr)","Your Model":kpis.get("recept_eui",0), "Good Practice":round(bm["good_eui"]*bm["recept_pct"]/100), "Median":round(bm["median_eui"]*bm["recept_pct"]/100), "High Flag":round(bm["high_eui"]*bm["recept_pct"]/100), "Status":status_color(kpis.get("recept_eui",0),round(bm["good_eui"]*bm["recept_pct"]/100),round(bm["median_eui"]*bm["recept_pct"]/100),round(bm["high_eui"]*bm["recept_pct"]/100))},
                {"Metric":"Pumps EUI (kWh/m²·yr)",     "Your Model":kpis["pumps_eui"],        "Good Practice":round(bm["good_eui"]*bm["pumps_pct"]/100),  "Median":round(bm["median_eui"]*bm["pumps_pct"]/100),  "High Flag":round(bm["high_eui"]*bm["pumps_pct"]/100),  "Status":status_color(kpis["pumps_eui"], round(bm["good_eui"]*bm["pumps_pct"]/100),  round(bm["median_eui"]*bm["pumps_pct"]/100),  round(bm["high_eui"]*bm["pumps_pct"]/100))},
                {"Metric":"GHGI (kgCO₂e/m²·yr)",       "Your Model":kpis["ghgi"],             "Good Practice":round(bm["median_ghgi"]*0.75),               "Median":bm["median_ghgi"],                            "High Flag":round(bm["median_ghgi"]*1.5),               "Status":status_color(kpis["ghgi"],round(bm["median_ghgi"]*0.75),bm["median_ghgi"],round(bm["median_ghgi"]*1.5))},
            ]
            st.dataframe(pd.DataFrame(rows),use_container_width=True,hide_index=True)

        st.divider()

        # ── Section 4: QA/QC Flags ──
        st.markdown("### 🚩 QA/QC Flags")
        st.caption("Automated checks comparing your model against benchmark thresholds.")
        for level,icon,msg in flags:
            st.markdown(f'<div class="flag-{level}"><b>{icon}</b> {msg}</div>',unsafe_allow_html=True)

        st.divider()

        # ── Export ──
        st.markdown("### 📥 Export")
        output=io.BytesIO()
        with pd.ExcelWriter(output,engine="openpyxl") as writer:
            pd.DataFrame({"Field":["Project Name","Building Type","City","Climate Zone","Software","Model Type","Phase","Date",
                "Floor Area (m²)","Total EUI","Electricity EUI","Gas EUI","Heating EUI","Cooling EUI","Fan EUI","Lighting EUI","DHW EUI","GHGI","Benchmark Percentile"],
                "Value":[meta["project_name"],meta["building_type"],meta["city"],meta["climate_zone"],meta["software"],meta["model_type"],meta["phase"],meta["date"],
                kpis["area"],kpis["total_eui"],kpis["elec_eui"],kpis["gas_eui"],kpis["heat_eui"],kpis["cool_eui"],kpis["fan_eui"],kpis["ltg_eui"],kpis["dhw_eui"],kpis["ghgi"],f"{pct}th" if pct else "N/A"],
            }).to_excel(writer,sheet_name="Summary",index=False)
            pd.DataFrame({"Level":[f[0].upper() for f in flags],"Icon":[f[1] for f in flags],"Message":[f[2] for f in flags]}).to_excel(writer,sheet_name="QA_QC_Flags",index=False)
            if bm: pd.DataFrame(rows).to_excel(writer,sheet_name="Benchmark",index=False)
        cd1,cd2=st.columns(2)
        with cd1:
            st.download_button("📥 Download Excel Report",data=output.getvalue(),
                file_name=f"QA_QC_{(meta['project_name'] or 'report').replace(' ','_')}_{meta['date']}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)
        with cd2:
            if st.button("← Run Another Project",use_container_width=True):
                for k in ["step","vals","results","headers","csv_df","mapping"]:
                    if k in st.session_state: del st.session_state[k]
                st.session_state.step=1; st.rerun()
